import speech_recognition as sr
import pyttsx3
import os
import webbrowser
import subprocess
import spotipy
import spotipy.util as util
import wolframalpha
import wikipedia
import datetime
import time
import openpyxl
import pandas as pd
import gui


engine = pyttsx3.init('sapi5')
voices = engine.getProperty('voices')
engine.setProperty('voice', voices[0].id)

chrome_path = "C:/Program Files/Google/Chrome/Application/chrome.exe"
excel_file_path = 'todos.xlsx'


def speak(text):
    engine.say(text)
    engine.runAndWait()


def takeCommand():
    r = sr.Recognizer()
    with sr.Microphone() as source:
        print("Vafs is Listening")
        r.adjust_for_ambient_noise(source)
        audio = r.listen(source)
    try:
        statement = r.recognize_google(audio, language='en-US')
        print(f"Your Message: {statement}\n")
        return statement.lower()

    except sr.UnknownValueError:
        speak("Sorry, I didn't understand what you said. Please try again.")
        return "None"
    except sr.RequestError:
        speak("Sorry, I couldn't connect to the speech recognition service. Please try again later.")
        return "None"


def ask_wolframalpha(question):
    client = wolframalpha.Client('W8EVX3-WR27HUJJHX')
    res = client.query(question)
    answer = next(res.results).text
    return answer


def play_spotify():
    # get Spotify access token
    scope = 'user-read-playback-state,user-modify-playback-state'
    token = util.prompt_for_user_token(username='31yqu26xpkrxri53njli2a4es6ae', scope=scope,
                                       client_id='1b325de7778a4d9baf5ea1205339c6b3',
                                       client_secret='c1f1281a89ff4cfc8ea0733ec4977bff',
                                       redirect_uri='http://localhost:8000/callback/')

    if token:
        sp = spotipy.Spotify(auth=token)

        # search for a song
        speak("What song would you like to play?")
        song_name = input("Please input here what song you want to play. ")
        results = sp.search(q=song_name, type='track', limit=1)

        # get the Spotify URI for the first search result
        uri = results['tracks']['items'][0]['uri']

        # play the song
        sp.start_playback(uris=[uri])

        # open the Spotify web player in a browser window
        webbrowser.open_new_tab("https://open.spotify.com/")
    else:
        print("Sorry, I couldn't authenticate with Spotify. Please try again later.")

def timegreet():
    now = datetime.datetime.now()
    hour = now.hour
    if hour < 12:
        speak("Good morning!")
    elif 12 <= hour < 18:
        speak("Good afternoon!")
    else:
        speak("Good evening!")

def standby_mode():
    speak("Going into standby mode")
    while True:
        command = takeCommand().lower()
        if command == "wake up":
            break
    speak("I'm back! How can I help you?")


def open_excel_file(file_path):
    os.startfile(file_path)


def display_to_do_list():
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active
        max_row = sheet.max_row

        if max_row <= 1:
            speak("Your to-do list is empty.")
        else:
            speak("Certainly. Please wait for a moment")
            time.sleep(1.5)
            speak("Your to do's are:")
            for row in range(2, max_row + 1):
                subject = sheet.cell(row=row, column=1).value
                description = sheet.cell(row=row, column=2).value
                due_date = sheet.cell(row=row, column=3).value

                speak(f"Subject: {subject}")
                speak(f"Description: {description}")
                speak(f"Due Date: {due_date}")

    except FileNotFoundError:
        speak("Your to-do list file was not found.")


def vafs():
    timegreet()
    speak("Hello! I am a voice assistant programmed to ease your work and help you. My name is Vafs. How can i help you today?")
    print("Welcome to Vafs! How can I help you?")
    standby_time = 10
    while True:
        command = takeCommand().lower()

        # open the Chrome browser
        if "open browser" in command or "open chrome" in command:
            subprocess.Popen([chrome_path])
            speak("Certainly")
            time.sleep(3)

        # open the spotify app
        elif "open spotify" in command or "play a song" in command:
            play_spotify()

        # give the to do s of the user
        elif "to do" in command or "homeworks" in command or "tasks" in command:
            display_to_do_list()

        # display the user's schedule
        elif 'schedule' in command:
            speak("Certainly, just wait for a moment")
            time.sleep(2)
            file_path = 'schedule.xlsx'
            open_excel_file(file_path)

        # open the notepad app
        elif "open notepad" in command:
            subprocess.Popen(["notepad.exe"])
            speak("Certainly")
            time.sleep(3)

        # open the YouTube tab in chrome
        elif "open youtube" in command or "open yt" in command:
            webbrowser.open("https://www.youtube.com")
            speak("Certainly")
            time.sleep(3)

        # open the Google Classroom tab
        elif "open my google classroom" in command or "open my classroom" in command:
            webbrowser.open_new_tab("https://classroom.google.com/u/1/h")
            speak("Certainly")
            time.sleep(3)

        # open the Google meet tab
        elif "open google docs" in command or "docs" in command:
            webbrowser.open_new_tab("https://docs.google.com/document/u/0/")
            speak("Certainly")
            time.sleep(3)

        # answer a question from wikipedia
        elif "wikipedia" in command:
            speak('Searching Wikipedia...')
            statement = command.replace("wikipedia", "")
            results = wikipedia.summary(command, sentences=3)
            speak("According to Wikipedia")
            print(results)
            speak(results)

        #
        elif "ask" in command or "question" in command:
            speak("Please input your question below.")
            question = input("What would you like to ask? ")
            answer = ask_wolframalpha(question)
            print(answer)
            speak(answer)

        # a standby feature
        elif "standby" in command or "sleep" in command:
            standby_mode_active = True
            standby_mode()
            standby_mode_active = False
        elif command == "Standby mode off":
            standby_mode_active = False
            speak("I'm back! How can I help you?")

        # shut the program down
        elif "bye" in command or "goodbye" in command or "exit" in command:
            speak("I am now shutting down. Goodbye.")
            break

         # q and a for the AI
        elif "tell me a joke" in command or "joke" in command:
            speak("Where did the software developer go?")
            time.sleep(2)
            speak("I don't know, maybe he ransomware")

        elif "what are you" in command or "who are you" in command:
            speak("I am Vafs, stands for Voice Assistant for Students")
            time.sleep(1)
            speak("I am a basic version of voice assistant create on fulfilling the final project of Dotes")
            time.sleep(0.5)
            speak("I am an AI voice assistant that can do some basic tasks like answering basic questions and opening various tabs and apps on your computer and to provide your tasks schedule in your classroom")


if __name__ == '__main__':
    vafs()


    gui.start_gui()

